import json
import re
from pathlib import Path

from openpyxl import load_workbook


CODE_SOURCE = Path("/Users/wiwynovina/Documents/makely 221 colors.xlsx")
OUTPUT = Path("src/data/makelyColors.ts")
LETTER_ORDER = "ABCDEFGHM"


def main() -> None:
    workbook = load_workbook(CODE_SOURCE, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    codes = []

    for row in sheet.iter_rows():
        for cell in row:
            value = str(cell.value).strip() if cell.value is not None else ""
            if re.fullmatch(r"[A-Z][0-9]{2}", value):
                codes.append(value)

    codes.sort(key=lambda code: (LETTER_ORDER.index(code[0]), int(code[1:])))

    existing = OUTPUT.read_text(encoding="utf-8")
    hex_values = re.findall(r'"hex": "(#[0-9A-F]{6})"', existing)
    if len(codes) != len(hex_values):
        raise ValueError(f"Expected matching code and hex counts, got {len(codes)} codes and {len(hex_values)} hex values")

    rows = [{"code": code, "name": code, "hex": hex_value} for code, hex_value in zip(codes, hex_values)]

    content = [
        "export const makelyColors = ",
        json.dumps(rows, indent=2),
        " as const;\n",
    ]
    OUTPUT.write_text("".join(content), encoding="utf-8")


if __name__ == "__main__":
    main()
